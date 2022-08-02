import sys
import datetime
import openpyxl

ExcelFile = "/Users/danielperron/Desktop/projet/SOLAR/Solar.xlsx"
solarFile = "/Users/danielperron/Desktop/projet/SOLAR/solar.txt"
#open excel
wb = openpyxl.load_workbook(ExcelFile)

#first vSys

format1 = 'yyyy\\-mm\\-dd'
format2 = 'hh:mm:ss'

now = datetime.datetime.now()


def TransferData(fileName,mask, ws, factor=1.0):
    f = open(fileName)
    count = 1
    while(True):
        l = f.readline()
        if len(l)==0 : break
        if l.find(mask) <0 :
             continue
        pos = l.find(".")
        thestamp = l[:pos].split(" ")
        thedate = [ int(i) for i in thestamp[0].split('-')]
        thetime = [ int(i) for i in thestamp[1].split(':')]
        exceldate = datetime.datetime(thedate[0],thedate[1],thedate[2],0,0)
        exceldatetime = datetime.datetime(thedate[0],thedate[1],thedate[2],thetime[0],thetime[1],thetime[2])
        ndays = (now - exceldatetime).days
        seconds = (now - exceldatetime).seconds
        #print(ndays)
        if ndays > 2:
            continue
#        if ndays > 1:
#           if seconds > (12 * 3600):
#             continue
            
        exceltime = datetime.time(thetime[0],thetime[1],thetime[2])
        pos = l.find('PicoW')
        topicvalue = ' '.join(l[pos:].split()).split(' ')
        value = float(topicvalue[1])

        ws.cell(row=count,column=1).value=exceldate
        ws.cell(row=count,column=1).number_format=format1
        ws.cell(row=count,column=2).value=exceltime
        ws.cell(row=count,column=2).number_format=format2
        ws.cell(row=count,column=3).value=exceldatetime
        ws.cell(row=count,column=3).number_format=format2
        ws.cell(row=count,column=4).value=value * factor
        count=count+1
    #ok now we need to erase everything left
    ws.delete_rows(count,10000)
    

TransferData(solarFile,"Vsys",wb['vSys'])
TransferData(solarFile,"Vsolar",wb['vSolar'],1.08)
TransferData(solarFile,"temp ",wb['ds18b20'])
TransferData(solarFile,"tempCPU",wb['cpu'])
TransferData(solarFile,"Vhydro",wb['vHydro'])
wb.save(ExcelFile)
wb.close()
